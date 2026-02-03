from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Font

# Standard colors from xlsx.md
COLOR_BLUE = "0000FF"


def _normalize_translated_text(block: dict, translated_text: str) -> str:
    if not translated_text:
        return ""
    if "\n" not in translated_text:
        return translated_text

    parts = [part.strip() for part in translated_text.splitlines() if part.strip()]
    if len(parts) <= 1:
        return parts[0] if parts else ""

    source_text = (block.get("source_text") or "").strip()
    source_norm = source_text.lower().strip()

    # 1) 若其中一行等於原文，移除後取剩下的最後一行
    if source_norm:
        filtered = [p for p in parts if p.lower().strip() != source_norm]
        if filtered:
            return filtered[-1]

    # 2) 若第一行包含原文，直接取最後一行
    if source_norm and source_norm in parts[0].lower():
        return parts[-1]

    # 3) 若可偵測語言，選與原文不同的行
    try:
        from backend.services.language_detect import detect_language

        source_lang = detect_language(source_text) if source_text else None
        if source_lang:
            candidates = [p for p in parts if detect_language(p) != source_lang]
            if candidates:
                return candidates[-1]
    except Exception:
        pass

    # 4) 回退：保留最後一行
    return parts[-1]


def _apply_translations_to_sheet(ws, translations: dict):
    for addr, text in translations.items():
        try:
            cell = ws[addr]

            # PROTECT FORMULAS: Do not overwrite if it's a formula
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue

            cell.value = text

            # Apply Blue color for hardcoded translations (per xlsx.md)
            if cell.font:
                new_font = copy(cell.font)
                new_font.color = COLOR_BLUE
                cell.font = new_font
            else:
                cell.font = Font(color=COLOR_BLUE)
        except Exception:
            continue


def _ensure_wrap_text(cell):
    if cell.alignment:
        new_alignment = copy(cell.alignment)
        new_alignment.wrapText = True
        cell.alignment = new_alignment
    else:
        cell.alignment = Alignment(wrapText=True)


def _make_translated_sheet_name(wb, sheet_name: str) -> str:
    base = f"{sheet_name}_翻譯"
    max_len = 31
    idx = 0
    while True:
        suffix = f"_{idx}" if idx > 0 else ""
        name = f"{base}{suffix}"
        if len(name) > max_len:
            name = f"{base[: max_len - len(suffix)]}{suffix}"
        if name not in wb.sheetnames:
            return name
        idx += 1


def apply_translations(input_path: str, output_path: str, blocks: list[dict]):
    """
    Apply translations to the Excel file.
    Follows xlsx.md: Blue text for hardcoded inputs/translations.
    """
    # Load with data_only=False to keep formulas
    wb = openpyxl.load_workbook(input_path, data_only=False)

    # Map (sheet_name, cell_address) -> translated_text.
    translations = {}
    for block in blocks:
        sheet_name = block.get("sheet_name")
        cell_address = block.get("cell_address")
        translated_text = _normalize_translated_text(
            block,
            block.get("translated_text", ""),
        )
        if sheet_name and cell_address:
            translations[(sheet_name, cell_address)] = translated_text

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_translations = {
            addr: text
            for (s_name, addr), text in translations.items()
            if s_name == sheet_name
        }
        _apply_translations_to_sheet(ws, sheet_translations)

    wb.save(output_path)

    # NEW: Trigger formula recalculation and error scanning
    from backend.services.xlsx.recalc import recalc_xlsx

    try:
        recalc_result = recalc_xlsx(output_path)
        # We could log this or return it to the caller if needed
        # For now, it ensures the file is updated by LibreOffice engine
    except Exception:
        pass


def apply_bilingual(
    input_path: str,
    output_path: str,
    blocks: list[dict],
    layout: str = "inline",
):
    """
    Apply bilingual translations back to the Excel file.
    Preserves original styles and optimizes alignment for multi-line content.
    """
    wb = openpyxl.load_workbook(input_path, data_only=False)

    translations = {}
    for block in blocks:
        sheet_name = block.get("sheet_name")
        cell_address = block.get("cell_address")
        source_text = block.get("source_text", "")
        translated_text = block.get("translated_text", "")
        if sheet_name and cell_address:
            # Simple bilingual: Source \n Translated
            translations[(sheet_name, cell_address)] = f"{source_text}\n{translated_text}"

    if layout in {"new_slide", "new_page", "new_sheet"}:
        translated_only = {}
        for block in blocks:
            sheet_name = block.get("sheet_name")
            cell_address = block.get("cell_address")
            translated_text = _normalize_translated_text(
                block,
                block.get("translated_text", ""),
            )
            if sheet_name and cell_address:
                translated_only[(sheet_name, cell_address)] = translated_text

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_ws = wb.copy_worksheet(ws)
            new_ws.title = _make_translated_sheet_name(wb, sheet_name)
            sheet_translations = {
                addr: text
                for (s_name, addr), text in translated_only.items()
                if s_name == sheet_name
            }
            _apply_translations_to_sheet(new_ws, sheet_translations)
        wb.save(output_path)
        from backend.services.xlsx.recalc import recalc_xlsx
        try:
            recalc_xlsx(output_path)
        except Exception:
            pass
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_translations = {
            addr: text
            for (s_name, addr), text in translations.items()
            if s_name == sheet_name
        }
        for addr, text in sheet_translations.items():
            try:
                cell = ws[addr]

                # PROTECT FORMULAS
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue

                cell.value = text

                # Apply Blue color to the whole cell for bilingual
                if cell.font:
                    new_font = copy(cell.font)
                    new_font.color = COLOR_BLUE
                    cell.font = new_font
                else:
                    cell.font = Font(color=COLOR_BLUE)

                # Ensure wrapText is True for multi-line bilingual content
                _ensure_wrap_text(cell)
            except Exception:
                continue

    wb.save(output_path)

    # NEW: Trigger formula recalculation and error scanning
    from backend.services.xlsx.recalc import recalc_xlsx

    try:
        recalc_xlsx(output_path)
    except Exception:
        pass
