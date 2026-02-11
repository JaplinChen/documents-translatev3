from __future__ import annotations

import zipfile
from typing import Any

import openpyxl

from backend.contracts import make_block
from backend.services.extract_utils import (
    is_exact_term_match,
    is_garbage_text,
    is_numeric_only,
    is_technical_terms_only,
    sanitize_extracted_text,
)
from backend.services.image_ocr import extract_image_text_blocks
from backend.services.language_detect import detect_document_languages
from backend.services.ocr_lang import resolve_ocr_lang_from_doc_lang


def extract_blocks(
    xlsx_path: str,
    preferred_lang: str | None = None,
    layout_params: dict[str, Any] | None = None,
) -> dict:
    """
    Extract text blocks from an Excel file.

    Returns standard block format compatible with the translation pipeline.
    Uses read_only mode for performance with large files.
    """
    # Use read_only=False if we need dimensions/hidden status reliably,
    # but read_only=True is much faster for large files.
    # To get hidden status, we must use read_only=False or accept it
    # might be missing.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    blocks: list[dict] = []

    params = layout_params or {}
    skip_numbers = params.get("skip_numbers", False)
    skip_dates = params.get("skip_dates", False) # Not implemented yet for Excel
    skip_code = params.get("skip_code", False)
    skip_translated = params.get("skip_translated", False) # Not implemented yet for Excel

    # Deduplication map: {text: first_block}
    dedup_map: dict[str, dict] = {}

    # Track metadata for the whole document
    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        is_sheet_hidden = ws.sheet_state != "visible"

        # Iterate through all cells that have values
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            # Check if row is hidden
            is_row_hidden = (
                ws.row_dimensions[row_idx].hidden if row_idx in ws.row_dimensions else False
            )

            for cell in row:
                if cell.value is None:
                    continue

                # Check if column is hidden
                col_letter = cell.column_letter
                is_col_hidden = (
                    ws.column_dimensions[col_letter].hidden
                    if col_letter in ws.column_dimensions
                    else False
                )

                # Skip common Excel error values
                cell_val_str = sanitize_extracted_text(str(cell.value))
                if cell_val_str.startswith("#") and any(
                    err in cell_val_str for err in ["REF!", "DIV/0!", "VALUE!", "N/A", "NAME?"]
                ):
                    continue

                text = cell_val_str

                # Use heuristics and layout_params to filter content
                if not text or is_garbage_text(text) or is_exact_term_match(text):
                    continue

                if is_numeric_only(text):
                    if skip_numbers:
                        continue
                    elif not params.get("force_extract_numbers"):
                        continue

                if skip_code and is_technical_terms_only(text):
                    continue

                # Location info for this occurrence
                location = {
                    "sheet_index": sheet_index,
                    "sheet_name": sheet_name,
                    "cell_address": cell.coordinate,
                    "is_hidden": is_sheet_hidden or is_row_hidden or is_col_hidden
                }

                if text in dedup_map:
                    # Already seen this text, check if this location is new to avoid duplicates
                    if not any(loc["sheet_index"] == sheet_index and loc["cell_address"] == cell.coordinate 
                               for loc in dedup_map[text]["locations"]):
                        dedup_map[text]["locations"].append(location)
                else:
                    # New unique text block
                    block = make_block(
                        slide_index=sheet_index,
                        shape_id=len(dedup_map) + 1,
                        block_type="spreadsheet_cell",
                        source_text=text,
                    )
                    # Custom properties for Excel
                    block["sheet_name"] = sheet_name
                    block["cell_address"] = cell.coordinate
                    block["locations"] = [location]
                    
                    # client_id is now based on content hash for stable tracking of deduplicated blocks
                    import hashlib
                    content_hash = hashlib.md5(text.encode("utf-8")).hexdigest()[:12]
                    block["client_id"] = f"xlsx-merged-{content_hash}"
                    
                    dedup_map[text] = block

    blocks = list(dedup_map.values())

    doc_lang = preferred_lang or (detect_document_languages(blocks).get("primary") if blocks else None)
    ocr_lang = resolve_ocr_lang_from_doc_lang(doc_lang)

    # Extract image text from embedded media
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]
            for idx, name in enumerate(media_files):
                try:
                    image_bytes = zf.read(name)
                except Exception:
                    continue
                blocks.extend(
                    extract_image_text_blocks(
                        image_bytes,
                        slide_index=-1,
                        shape_id=f"xlsx-image-{idx}",
                        image_part=name,
                        source="xlsx",
                        ocr_lang=ocr_lang,
                    )
                )
    except Exception:
        pass

    return {
        "blocks": blocks,
        "sheet_count": len(wb.sheetnames),
    }
