import os
import subprocess
import json
import sys
from typing import Optional


def recalc_xlsx(file_path: str, timeout_seconds: int = 30) -> dict:
    """
    Recalculate formulas in an XLSX file using LibreOffice.
    Scans for Excel errors after recalculation.
    """
    if not os.path.exists(file_path):
        return {"status": "error", "message": f"File not found: {file_path}"}

    # Use soffice to recalculate and save
    # Command: soffice --headless --convert-to xlsx --outdir <dir> <file>
    # Note: LibreOffice automatically recalculates on open/save if configured.
    # To force recalculation, we can use a macro or just the conversion command
    # which usually triggers a recalc of volatile functions at least.

    # However, standard openpyxl doesn't evaluate formulas.
    # The 'recalc.py' mentioned in xlsx.md suggests a more robust approach.
    # Here we will attempt a best-effort recalc using headless LibreOffice.

    try:
        # We use --convert-to xlsx to the same file to trigger LibreOffice's engine
        out_dir = os.path.dirname(file_path)
        cmd = ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, file_path]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout_seconds)

        if result.returncode != 0:
            return {
                "status": "error",
                "message": f"LibreOffice failed: {result.stderr}",
                "libreoffice_installed": False,
            }

    except subprocess.TimeoutExpired:
        return {"status": "error", "message": "LibreOffice timeout"}
    except FileNotFoundError:
        return {
            "status": "error",
            "message": "LibreOffice (soffice) not found in PATH",
            "libreoffice_installed": False,
        }

    # Scan for errors using openpyxl in data_only mode
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        error_summary = {}
        total_errors = 0
        total_formulas = 0

        error_types = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val_str = str(cell.value)
                    if any(err in val_str for err in error_types):
                        total_errors += 1
                        if val_str not in error_summary:
                            error_summary[val_str] = {"count": 0, "locations": []}
                        error_summary[val_str]["count"] += 1
                        error_summary[val_str]["locations"].append(
                            f"{sheet_name}!{cell.coordinate}"
                        )

        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": error_summary,
            "libreoffice_installed": True,
        }
    except Exception as e:
        return {"status": "error", "message": f"Error scanning for Excel errors: {str(e)}"}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "No file path provided"}))
        sys.exit(1)

    path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc_xlsx(path, timeout), indent=2))
