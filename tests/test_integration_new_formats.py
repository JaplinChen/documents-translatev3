import json
import os

import anyio
import httpx
import openpyxl
import pytest
from httpx import ASGITransport
from reportlab.pdfgen import canvas

from backend.main import app

class SyncASGIClient:
    def request(self, method: str, url: str, **kwargs):
        async def _run():
            transport = ASGITransport(app=app, raise_app_exceptions=True)
            async with httpx.AsyncClient(
                transport=transport, base_url="http://test"
            ) as client:
                return await client.request(method, url, **kwargs)

        return anyio.run(_run)

    def get(self, url: str, **kwargs):
        return self.request("GET", url, **kwargs)

    def post(self, url: str, **kwargs):
        return self.request("POST", url, **kwargs)


@pytest.fixture
def client():
    return SyncASGIClient()

@pytest.fixture
def dummy_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Hello World"
    ws["A2"] = 123  # Should be ignored by extractor (numeric)
    ws["B1"] = "Translation Test" # Plain string
    # Add some style
    ws["A1"].font = openpyxl.styles.Font(bold=True, color="FF0000")

    file_path = "test_input.xlsx"
    wb.save(file_path)
    yield file_path
    if os.path.exists(file_path):
        os.remove(file_path)

@pytest.fixture
def dummy_pdf():
    file_path = "test_input.pdf"
    c = canvas.Canvas(file_path)
    c.drawString(100, 750, "This is a test PDF content.")
    c.save()
    yield file_path
    if os.path.exists(file_path):
        os.remove(file_path)

def test_xlsx_integration(dummy_xlsx, client):
    # 1. Test Extract
    with open(dummy_xlsx, "rb") as f:
        response = client.post(
            "/api/xlsx/extract",
            files={
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    data = response.json()
    assert len(data["blocks"]) >= 2
    # Check if texts are extracted
    source_texts = [b["source_text"] for b in data["blocks"]]
    assert "Hello World" in source_texts
    assert "Translation Test" in source_texts

    # 2. Test Apply
    blocks = data["blocks"]
    for b in blocks:
        b["translated_text"] = "你好世界"

    with open(dummy_xlsx, "rb") as f:
        response = client.post(
            "/api/xlsx/apply",
            data={
                "blocks": json.dumps(blocks),
                "mode": "translated",
            },
            files={
                "file": (
                    "test.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    assert "download_url" in response.json()
    assert response.json()["filename"].endswith(".xlsx")

def test_pdf_integration(dummy_pdf, client):
    # 1. Test Extract
    with open(dummy_pdf, "rb") as f:
        response = client.post(
            "/api/pdf/extract",
            files={"file": ("test.pdf", f, "application/pdf")},
        )

    assert response.status_code == 200
    data = response.json()
    assert len(data["blocks"]) >= 1
    assert "This is a test PDF content." in [b["source_text"] for b in data["blocks"]]

    # 2. Test Apply
    blocks = data["blocks"]
    for b in blocks:
        b["translated_text"] = "這是 PDF 測試內容"

    with open(dummy_pdf, "rb") as f:
        response = client.post("/api/pdf/apply", data={
            "blocks": json.dumps(blocks),
            "mode": "translated"
        }, files={"file": ("test.pdf", f, "application/pdf")})

    assert response.status_code == 200
    assert "download_url" in response.json()
    assert response.json()["filename"].endswith(".pdf")
