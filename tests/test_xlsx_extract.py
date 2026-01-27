import os

import pytest
from openpyxl import Workbook

from backend.services.xlsx.extract import extract_blocks

def test_xlsx_extraction(tmp_path):
    # Create a dummy XLSX file
    xlsx_path = os.path.join(tmp_path, "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello World"
    ws["B2"] = "12345" # Should be filtered (numeric)
    ws["C3"] = "API_KEY" # Should be filtered (technical)
    ws["D4"] = "這是一個測試"
    wb.save(xlsx_path)

    result = extract_blocks(xlsx_path)
    blocks = result["blocks"]

    # Check if only meaningful text is extracted
    texts = [b["source_text"] for b in blocks]
    assert "Hello World" in texts
    assert "這是一個測試" in texts
    assert "12345" not in texts
    assert len(blocks) == 2
    assert blocks[0]["sheet_name"] == wb.active.title
    assert blocks[0]["cell_address"] == "A1"

if __name__ == "__main__":
    pytest.main([__file__])
