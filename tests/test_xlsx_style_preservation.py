import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.services.xlsx.apply import apply_bilingual, apply_translations

def test_xlsx_style_deep_preservation():
    # 1. Create a complex styled XLSX
    input_path = "style_test_in.xlsx"
    output_path = "style_test_out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["B2"]
    cell.value = "Original Text"

    # Apply complex styles
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    cell.fill = blue_fill
    cell.border = thin_border
    cell.font = Font(bold=True, size=14)
    cell.number_format = "@" # Text format
    cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(input_path)

    # 2. Run apply_translations
    blocks = [{
        "sheet_name": "Sheet",
        "cell_address": "B2",
        "translated_text": "Translated Text"
    }]

    apply_translations(input_path, output_path, blocks)

    # 3. Verify
    wb_out = openpyxl.load_workbook(output_path)
    cell_out = wb_out["Sheet"]["B2"]

    assert cell_out.value == "Translated Text"
    assert cell_out.fill.start_color.rgb == "000000FF" # openpyxl adds Alpha
    assert cell_out.border.left.style == "thin"
    assert cell_out.font.bold is True
    assert cell_out.number_format == "@"

    os.remove(input_path)
    os.remove(output_path)

def test_xlsx_bilingual_style_preservation():
    input_path = "bilingual_style_test_in.xlsx"
    output_path = "bilingual_style_test_out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["C3"]
    cell.value = "Source"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.fill = red_fill
    wb.save(input_path)

    blocks = [{
        "sheet_name": "Sheet",
        "cell_address": "C3",
        "source_text": "Source",
        "translated_text": "Target"
    }]

    apply_bilingual(input_path, output_path, blocks)

    wb_out = openpyxl.load_workbook(output_path)
    cell_out = wb_out["Sheet"]["C3"]

    assert "Source" in cell_out.value
    assert "Target" in cell_out.value
    # openpyxl normalized color check
    assert cell_out.fill.start_color.rgb in ["00FF0000", "FFFF0000"]
    # Bilingual should have wrapText enabled
    assert cell_out.alignment.wrapText is True

    os.remove(input_path)
    os.remove(output_path)
